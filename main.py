import openpyxl
import os
from pathlib import Path

d = dict([])

with open("conf.txt", "r") as f:
    for line in f:
       d[line.partition('=')[0].strip()] = line.partition('=')[2].strip()

def getdata(list_path, path_dir):
    
    print('Путь к файлу получен успешно')

    sheet = openpyxl.load_workbook(list_path).active  #открываем перечень КИП и переходим на первую страницу
    print('Перечень КИП открыт')
    for row in range(3, sheet.max_row + 1):
        massive = [0] * sheet.max_column
        print('Буфер очищен успешно')
        for col in range(0, sheet.max_column):
            massive[col] = sheet[row][col].value
            print('Считано значение строки ', row, ' столбца ', col, ' содержащее ', massive[col])
        Designator = massive[6]
        Pos = massive[7] 
        book = openpyxl.load_workbook(d[Designator])  #Открываем файл шаблона
        sh = book.active
        print('Шаблон открыт успешно')

        conf_path = d[Designator].partition('.')[0].strip() + ".txt"

        with open(conf_path, "r") as f:
            for line in f:
                sh[line.partition('=')[0].strip()].value = massive[int(line.partition('=')[2].strip())]
                print('В ячейку шаблона ', line.partition('=')[0].strip(), 'записано', sh[line.partition('=')[0].strip()].value)
        
        file_name = str(Designator) + ' ' + str(Pos) + '.xlsx'
        
        file_name = file_name.replace('/', '-')
        
        dubbed_file_name = 'Дублированный ' + file_name
         
       
        if not os.path.exists(file_name):
            book.save(file_name)
            print('Опросный лист на датчик ' + file_name + ' создан успешно!')
        else:
            print('Опросный лист на датчик ' + file_name + ' существует. Заменить опросный лист с повторяющейся позицией? (Да/Нет)')
            allow = input()
            if allow == 'Да' or allow == 'да' or allow == 'ДА' or allow == 'дА':
                book.save(file_name)
                print('Опросный лист на датчик ' + file_name + ' заменен!')
            else:
                book.save(dubbed_file_name)
                print('Создан опросный лист на датчик с повторяющейся позицией ' + str(dubbed_file_name))

        book.close()
    
    return(print('Создание опросных листов завершено успешно!'))    

print('Задайте путь к Перечню КИП')

list_path = input()

print('Куда сохранить?')

path_dir = Path(input())

getdata(list_path, path_dir)
